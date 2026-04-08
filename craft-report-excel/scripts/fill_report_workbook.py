from __future__ import annotations

import argparse
import json
import shutil
from copy import copy
from datetime import date, datetime, timedelta
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
ASSETS_DIR = SKILL_DIR / "assets"
DEFAULT_TEMPLATE = ASSETS_DIR / "craft_report_template.xlsx"
REFERENCE_TEMPLATE = ASSETS_DIR / "工艺工程与制造部每日总结-open claw用.xlsx"

UNIT_ROWS = {
    "光刻": (2, 3),
    "刻蚀": (4, 5),
    "薄膜": (6, 7),
    "掺杂": (8, 9),
    "晶背": (10, 11),
    "检测": (12, 13),
    "生产": (14, 15),
    "整合": (16, 17),
}

VALID_COLUMNS = ("B", "C", "D", "E", "F", "G")
WEEKLY_COLUMNS = ("B", "G")
DAILY_COLUMNS = ("C", "D", "E", "F")
WEEKDAY_CN = {
    0: "周一",
    1: "周二",
    2: "周三",
    3: "周四",
    4: "周五",
    5: "周六",
    6: "周日",
}

FALLBACK_CELL_LABELS = {
    "B": {
        "upper": "本周总结：",
        "lower": "下周计划（如周六、周日计划加班，工作计划需在此处一并体现）：",
    },
    "C": {
        "upper": "今日总结（如周六、周日有加班，工作总结需在此处体现）",
        "lower": "次日计划：",
    },
    "D": {
        "upper": "今日总结：",
        "lower": "次日计划：",
    },
    "E": {
        "upper": "今日总结：",
        "lower": "次日计划：",
    },
    "F": {
        "upper": "今日总结：",
        "lower": "次日计划：",
    },
    "G": {
        "upper": "本周总结：",
        "lower": "下周计划（如周六、周日计划加班，工作计划需在此处一并体现）：",
    },
}


def load_sheet(path: Path):
    wb = load_workbook(path)
    return wb, wb[wb.sheetnames[0]]


def normalize_multiline_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return text


def default_label_for_cell(column: str, position: str) -> str:
    return FALLBACK_CELL_LABELS[column][position]


def is_label_like(text: str) -> bool:
    return ("总结" in text) or ("计划" in text)


@lru_cache(maxsize=1)
def load_reference_cell_labels() -> dict[str, dict[str, str]]:
    labels: dict[str, dict[str, str]] = {}

    if not REFERENCE_TEMPLATE.exists():
        return labels

    _, ws = load_sheet(REFERENCE_TEMPLATE)
    for _, (upper_row, lower_row) in UNIT_ROWS.items():
        for column, position, row in (
            *((col, "upper", upper_row) for col in VALID_COLUMNS),
            *((col, "lower", lower_row) for col in VALID_COLUMNS),
        ):
            text = normalize_multiline_text(ws[f"{column}{row}"].value)
            first_line = text.split("\n", 1)[0] if text else ""
            if is_label_like(first_line):
                labels.setdefault(column, {})[position] = first_line

    return labels


def label_for_cell(column: str, position: str) -> str:
    return load_reference_cell_labels().get(column, {}).get(position, default_label_for_cell(column, position))


def ensure_wrapped_cell(ws, coordinate: str) -> None:
    cell = ws[coordinate]
    alignment = copy(cell.alignment)
    alignment.wrap_text = True
    cell.alignment = alignment


def format_labeled_value(column: str, position: str, content: object) -> str:
    label = label_for_cell(column, position)
    text = normalize_multiline_text(content)

    if not text:
        return label

    if text.startswith(label):
        return text

    return f"{label}\n{text}"


def apply_default_body_labels(ws) -> None:
    for upper_row, lower_row in UNIT_ROWS.values():
        for col in VALID_COLUMNS:
            ws[f"{col}{upper_row}"] = label_for_cell(col, "upper")
            ws[f"{col}{lower_row}"] = label_for_cell(col, "lower")
            ensure_wrapped_cell(ws, f"{col}{upper_row}")
            ensure_wrapped_cell(ws, f"{col}{lower_row}")


def resolve_source(source: Path | None) -> Path:
    if source:
        return source
    if not DEFAULT_TEMPLATE.exists():
        raise FileNotFoundError(f"Default template not found: {DEFAULT_TEMPLATE}")
    return DEFAULT_TEMPLATE


def parse_friday(value: str) -> date:
    friday = datetime.strptime(value, "%Y-%m-%d").date()
    if friday.weekday() != 4:
        raise ValueError(f"{value} is not a Friday.")
    return friday


def format_cn_date(value: date) -> str:
    return f"{value.year}/{value.month}/{value.day}（{WEEKDAY_CN[value.weekday()]}）"


def build_headers_for_friday(friday_value: str) -> dict[str, str]:
    friday = parse_friday(friday_value)
    monday = friday - timedelta(days=4)
    return {
        "B": format_cn_date(friday - timedelta(days=7)),
        "C": format_cn_date(monday),
        "D": format_cn_date(monday + timedelta(days=1)),
        "E": format_cn_date(monday + timedelta(days=2)),
        "F": format_cn_date(monday + timedelta(days=3)),
        "G": format_cn_date(friday),
    }


def print_headers_for_friday(friday_value: str) -> None:
    print(json.dumps(build_headers_for_friday(friday_value), ensure_ascii=False, indent=2))


def dump_workbook(source: Path) -> None:
    _, ws = load_sheet(source)
    data = {
        "sheet": ws.title,
        "headers": {col: ws[f"{col}1"].value for col in VALID_COLUMNS},
        "units": {},
    }

    for unit, (upper_row, lower_row) in UNIT_ROWS.items():
        data["units"][unit] = {}
        for col in VALID_COLUMNS:
            data["units"][unit][col] = {
                "upper": ws[f"{col}{upper_row}"].value,
                "lower": ws[f"{col}{lower_row}"].value,
            }

    print(json.dumps(data, ensure_ascii=False, indent=2))


def clear_targets(ws, headers: dict[str, str] | None, units: dict[str, dict] | None) -> None:
    if headers:
        for col in headers:
            ws[f"{col}1"] = None

    if units:
        for unit, column_map in units.items():
            upper_row, lower_row = UNIT_ROWS[unit]
            for col in column_map:
                ws[f"{col}{upper_row}"] = label_for_cell(col, "upper")
                ws[f"{col}{lower_row}"] = label_for_cell(col, "lower")
                ensure_wrapped_cell(ws, f"{col}{upper_row}")
                ensure_wrapped_cell(ws, f"{col}{lower_row}")


def clear_all_report_content(ws) -> None:
    for col in VALID_COLUMNS:
        ws[f"{col}1"] = None

    apply_default_body_labels(ws)


def apply_fixed_headers(ws, friday_value: str | None, explicit_headers: dict[str, str]) -> dict[str, str]:
    headers = dict(explicit_headers)
    if friday_value:
        headers.update(build_headers_for_friday(friday_value))
    for col, value in headers.items():
        ws[f"{col}1"] = value
        ensure_wrapped_cell(ws, f"{col}1")
    return headers


def prepare_output_from_source(source: Path, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.resolve() != source.resolve() or not output.exists():
        shutil.copy2(source, output)


def create_blank_workbook(source: Path | None, output: Path, friday_value: str | None) -> None:
    actual_source = resolve_source(source)
    prepare_output_from_source(actual_source, output)
    wb, ws = load_sheet(output)
    clear_all_report_content(ws)
    apply_fixed_headers(ws, friday_value, {})
    wb.save(output)


def fill_workbook(
    source: Path | None,
    output: Path,
    payload_path: Path,
    clear_target_cells: bool,
    clear_all_content: bool,
    friday_value: str | None,
) -> None:
    with payload_path.open("r", encoding="utf-8-sig") as f:
        payload = json.load(f)

    actual_source = resolve_source(source)
    prepare_output_from_source(actual_source, output)
    wb, ws = load_sheet(output)

    headers = payload.get("headers", {})
    units = payload.get("units", {})

    validate_payload(headers, units)

    if clear_all_content:
        clear_all_report_content(ws)
    elif clear_target_cells:
        clear_targets(ws, headers, units)

    apply_fixed_headers(ws, friday_value, headers)

    for unit, column_map in units.items():
        upper_row, lower_row = UNIT_ROWS[unit]
        for col, cell_pair in column_map.items():
            if "upper" in cell_pair:
                ws[f"{col}{upper_row}"] = format_labeled_value(col, "upper", cell_pair["upper"])
                ensure_wrapped_cell(ws, f"{col}{upper_row}")
            if "lower" in cell_pair:
                ws[f"{col}{lower_row}"] = format_labeled_value(col, "lower", cell_pair["lower"])
                ensure_wrapped_cell(ws, f"{col}{lower_row}")

    wb.save(output)


def validate_payload(headers: dict[str, str], units: dict[str, dict]) -> None:
    invalid_headers = [col for col in headers if col not in VALID_COLUMNS]
    if invalid_headers:
        raise ValueError(f"Unsupported header columns: {invalid_headers}")

    invalid_units = [unit for unit in units if unit not in UNIT_ROWS]
    if invalid_units:
        raise ValueError(f"Unsupported units: {invalid_units}")

    for unit, column_map in units.items():
        invalid_cols = [col for col in column_map if col not in VALID_COLUMNS]
        if invalid_cols:
            raise ValueError(f"Unsupported columns for {unit}: {invalid_cols}")
        for col, cell_pair in column_map.items():
            if not isinstance(cell_pair, dict):
                raise ValueError(f"Cell payload for {unit} {col} must be an object")
            invalid_keys = [key for key in cell_pair if key not in ("upper", "lower")]
            if invalid_keys:
                raise ValueError(f"Unsupported keys for {unit} {col}: {invalid_keys}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Inspect, initialize, or fill the craft department Excel report workbook."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    dump_parser = subparsers.add_parser("dump", help="Print workbook headers and unit cell text as JSON.")
    dump_parser.add_argument("--source", type=Path, help="Workbook to inspect. Defaults to the forced template.")

    headers_parser = subparsers.add_parser(
        "headers-for-friday",
        help="Generate the fixed B:G headers from a target Friday date.",
    )
    headers_parser.add_argument("--friday", required=True, help="Friday in YYYY-MM-DD format.")

    init_parser = subparsers.add_parser(
        "init-blank",
        help="Create a blank workbook from the forced template, preserving format but clearing all report content.",
    )
    init_parser.add_argument("--source", type=Path, help="Optional source workbook path. Defaults to the forced template.")
    init_parser.add_argument("--output", required=True, type=Path, help="Blank workbook output path.")
    init_parser.add_argument("--friday", help="Friday in YYYY-MM-DD format for fixed B:G headers.")

    fill_parser = subparsers.add_parser("fill", help="Fill workbook headers and unit cells from a JSON payload.")
    fill_parser.add_argument("--source", type=Path, help="Optional source workbook path. Defaults to the forced template.")
    fill_parser.add_argument("--output", required=True, type=Path, help="Filled workbook output path.")
    fill_parser.add_argument("--payload", required=True, type=Path, help="JSON payload file.")
    fill_parser.add_argument("--friday", help="Friday in YYYY-MM-DD format for fixed B:G headers.")
    fill_parser.add_argument(
        "--clear-targets",
        action="store_true",
        help="Clear target headers and cells before writing new values.",
    )
    fill_parser.add_argument(
        "--clear-all-content",
        action="store_true",
        help="Clear all report headers and report-body cells before writing. Use this for new workbook creation.",
    )

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    if args.command == "dump":
        dump_workbook(resolve_source(args.source))
        return

    if args.command == "headers-for-friday":
        print_headers_for_friday(args.friday)
        return

    if args.command == "init-blank":
        create_blank_workbook(args.source, args.output, args.friday)
        return

    if args.command == "fill":
        fill_workbook(
            args.source,
            args.output,
            args.payload,
            args.clear_targets,
            args.clear_all_content,
            args.friday,
        )
        return

    raise ValueError(f"Unsupported command: {args.command}")


if __name__ == "__main__":
    main()
