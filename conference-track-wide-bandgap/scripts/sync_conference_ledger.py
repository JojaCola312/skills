from __future__ import annotations

import argparse
import re
from datetime import date
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


SUMMARY_SHEET = "汇总信息"
MAIN_SHEET = "国内会议"
TRACKING_SHEET = "持续跟踪名单"
AUDIT_SHEET = "种子池核验清单"
NOTES_SHEET = "检索说明"

SUMMARY_SHEET_ALIASES = {SUMMARY_SHEET, "姹囨€讳俊鎭?"}
MAIN_SHEET_ALIASES = {MAIN_SHEET, "鍥藉唴浼氳"}
TRACKING_SHEET_ALIASES = {TRACKING_SHEET, "鎸佺画璺熻釜鍚嶅崟"}
AUDIT_SHEET_ALIASES = {AUDIT_SHEET, "绉嶅瓙姹犳牳楠屾竻鍗?"}
NOTES_SHEET_ALIASES = {NOTES_SHEET, "妫€绱㈣鏄?"}

MAIN_HEADERS = [
    "会议名称",
    "会议主题",
    "届次/年份",
    "起止日期",
    "会议地点",
    "会议类型",
    "官方网站/报名链接",
    "主办单位",
    "信息来源",
    "与半导体相关性",
    "最新状态",
    "最后核验",
]
MAIN_HEADER_ALIASES = {
    "会议名称": {"会议名称", "浼氳鍚嶇О"},
    "会议主题": {"会议主题", "浼氳涓婚"},
    "届次/年份": {"届次/年份", "灞婃/骞翠唤"},
    "起止日期": {"起止日期", "璧锋鏃ユ湡"},
    "会议地点": {"会议地点", "浼氳鍦扮偣"},
    "会议类型": {"会议类型", "浼氳绫诲瀷"},
    "官方网站/报名链接": {"官方网站/报名链接", "瀹樻柟缃戠珯/鎶ュ悕閾炬帴"},
    "主办单位": {"主办单位", "涓诲姙鍗曚綅", "主办单位/承办单位", "涓诲姙鍗曚綅/鎵垮姙鍗曚綅"},
    "信息来源": {"信息来源", "淇℃伅鏉ユ簮", "信息来源渠道", "淇℃伅鏉ユ簮娓犻亾"},
    "与半导体相关性": {"与半导体相关性", "涓庡崐瀵间綋鐩稿叧鎬?", "涓庡崐瀵间綋鐩稿叧鎬"},
    "最新状态": {"最新状态", "鏈€鏂扮姸鎬?"},
    "最后核验": {"最后核验", "鏈€鍚庢牳楠?", "鏈€鍚庢牳楠宍"},
}

TRACKING_HEADERS = [
    "会议名称",
    "上一届信息",
    "本轮核验结果",
    "待跟踪点",
    "优先检查网址",
]
TRACKING_HEADER_ALIASES = {
    "会议名称": {"会议名称", "浼氳鍚嶇О"},
    "上一届信息": {"上一届信息", "涓婁竴灞婁俊鎭?"},
    "本轮核验结果": {"本轮核验结果", "鏈疆鏍搁獙缁撴灉"},
    "待跟踪点": {"待跟踪点", "寰呰窡韪偣"},
    "优先检查网址": {"优先检查网址", "浼樺厛妫€鏌ョ綉鍧€"},
}

AUDIT_HEADERS = [
    "标准名",
    "是否有优先来源",
    "本轮搜索状态",
    "本轮使用关键词",
    "结果去向",
]
AUDIT_HEADER_ALIASES = {
    "标准名": {"标准名", "鏍囧噯鍚?"},
    "是否有优先来源": {"是否有优先来源", "鏄惁鏈変紭鍏堟潵婧?"},
    "本轮搜索状态": {"本轮搜索状态", "鏈疆鎼滅储鐘舵€?"},
    "本轮使用关键词": {"本轮使用关键词", "鏈疆浣跨敤鍏抽敭璇?"},
    "结果去向": {"结果去向", "缁撴灉鍘诲悜"},
}

SUMMARY_ALLOWED_KEYS = {
    "Report Title": {"Report Title"},
    "核验日期": {"核验日期", "鏍搁獙鏃ユ湡"},
    "种子池家族总数": {"种子池家族总数", "绉嶅瓙姹犲鏃忔€绘暟"},
    "本轮已核验": {"本轮已核验", "鏈疆宸叉牳楠?"},
    "确认国内会议": {"确认国内会议", "纭鍥藉唴浼氳"},
    "持续跟踪": {"持续跟踪", "鎸佺画璺熻釜"},
}


def sanitize(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def normalize_sheet_name(name: str) -> str:
    if name in SUMMARY_SHEET_ALIASES:
        return SUMMARY_SHEET
    if name in MAIN_SHEET_ALIASES:
        return MAIN_SHEET
    if name in TRACKING_SHEET_ALIASES:
        return TRACKING_SHEET
    if name in AUDIT_SHEET_ALIASES:
        return AUDIT_SHEET
    if name in NOTES_SHEET_ALIASES:
        return NOTES_SHEET
    return name


def parse_key_value_cell(text: str) -> Tuple[str, str]:
    if "：" in text:
        key, value = text.split("：", 1)
        return key.strip(), value.strip()
    if ":" in text:
        key, value = text.split(":", 1)
        return key.strip(), value.strip()
    return text.strip(), ""


def canonical_key(value: str, aliases: Dict[str, Iterable[str]]) -> str:
    text = sanitize(value)
    for canonical, candidates in aliases.items():
        if text in candidates:
            return canonical
    return text


def get_sheet(wb, aliases: Iterable[str]):
    for name in aliases:
        if name in wb.sheetnames:
            return wb[name]
    raise KeyError(f"Workbook missing required sheet. Expected one of: {sorted(set(aliases))}")


def read_summary_sheet(ws) -> Dict[str, str]:
    result: Dict[str, str] = {}
    title = sanitize(ws.cell(1, 1).value)
    if title:
        result["Report Title"] = title
    for row in ws.iter_rows(values_only=True):
        cell = sanitize(row[0]) if row and row[0] is not None else ""
        if not cell or cell == title:
            continue
        key, value = parse_key_value_cell(cell)
        key = canonical_key(key, SUMMARY_ALLOWED_KEYS)
        if key in SUMMARY_ALLOWED_KEYS:
            result[key] = value
    return result


def normalize_headers(actual_headers: List[str], aliases: Dict[str, Iterable[str]], expected: List[str]) -> List[str]:
    normalized = [canonical_key(header, aliases) for header in actual_headers]
    if normalized != expected:
        raise ValueError(f"Sheet headers do not match expected template. Got: {normalized}")
    return normalized


def read_table_sheet(ws, headers: List[str], aliases: Dict[str, Iterable[str]]) -> List[Dict[str, str]]:
    actual_headers = [sanitize(c.value) for c in ws[1][: len(headers)]]
    normalize_headers(actual_headers, aliases, headers)
    rows: List[Dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = [sanitize(v) for v in row[: len(headers)]]
        if not any(values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def read_notes_sheet(ws) -> List[Tuple[str, str]]:
    notes: List[Tuple[str, str]] = []
    for row in ws.iter_rows(values_only=True):
        left = sanitize(row[0]) if len(row) >= 1 else ""
        right = sanitize(row[1]) if len(row) >= 2 else ""
        if left or right:
            notes.append((left, right))
    return notes


def workbook_to_data(xlsx_path: Path) -> Dict[str, object]:
    wb = load_workbook(xlsx_path)
    return {
        "summary": read_summary_sheet(get_sheet(wb, SUMMARY_SHEET_ALIASES)),
        "main_pool": read_table_sheet(get_sheet(wb, MAIN_SHEET_ALIASES), MAIN_HEADERS, MAIN_HEADER_ALIASES),
        "tracking_list": read_table_sheet(get_sheet(wb, TRACKING_SHEET_ALIASES), TRACKING_HEADERS, TRACKING_HEADER_ALIASES),
        "seed_audit": read_table_sheet(get_sheet(wb, AUDIT_SHEET_ALIASES), AUDIT_HEADERS, AUDIT_HEADER_ALIASES)
        if any(name in wb.sheetnames for name in AUDIT_SHEET_ALIASES)
        else [],
        "search_notes": read_notes_sheet(get_sheet(wb, NOTES_SHEET_ALIASES))
        if any(name in wb.sheetnames for name in NOTES_SHEET_ALIASES)
        else [],
    }


def dump_record_block(lines: List[str], title: str, record: Dict[str, str], ordered_keys: List[str]) -> None:
    lines.append(f"### {title}")
    for key in ordered_keys:
        lines.append(f"- {key}: {sanitize(record.get(key, ''))}")
    hint = sanitize(record.get("Retrieval Hint", ""))
    if hint:
        lines.append(f"- Retrieval Hint: {hint}")
    lines.append("")


def data_to_markdown(data: Dict[str, object]) -> str:
    lines: List[str] = [
        "# Conference Ledger",
        "",
        "This file is the canonical ledger. Update this Markdown first, then regenerate Excel from it.",
        "",
        "## Ledger Rules",
        "",
        "- Treat this file as the source of truth. Do not update Excel first.",
        "- Keep section names unchanged: `Summary`, `Main Pool`, `Tracking List`, `Search Notes`.",
        "- Keep one `###` block per conference entry. Do not merge multiple conferences into one block.",
        "- In `Main Pool`, keep only sufficiently verified China-based events.",
        "- In `Tracking List`, keep unresolved, weak-evidence, or next-edition-unconfirmed events.",
        "- Prefer stable official or organizer links in `官方网站/报名链接`. If unavailable, keep `待确认` and record weak sources under `信息来源`.",
        "- Use `Retrieval Hint` to record how the event was found last time.",
        "",
        "## Summary",
        "",
    ]
    summary_order = ["Report Title", "核验日期", "种子池家族总数", "本轮已核验", "确认国内会议", "持续跟踪"]
    summary = data.get("summary", {})
    for key in summary_order:
        lines.append(f"- {key}: {sanitize(summary.get(key, ''))}")

    lines.extend(["", "## Main Pool", ""])
    for record in data.get("main_pool", []):
        dump_record_block(lines, record.get("会议名称", "Untitled Conference"), record, MAIN_HEADERS)

    lines.extend(["## Tracking List", ""])
    for record in data.get("tracking_list", []):
        dump_record_block(lines, record.get("会议名称", "Untitled Tracking Conference"), record, TRACKING_HEADERS)

    lines.extend(["## Search Notes", ""])
    for left, right in data.get("search_notes", []):
        if right:
            lines.append(f"- {left}: {right}")
        else:
            lines.append(f"- {left}")
    lines.append("")
    return "\n".join(lines)


def audit_to_markdown(data: Dict[str, object]) -> str:
    lines: List[str] = [
        "# Seed Audit",
        "",
        "This file stores per-run seed-family verification state. It is operational state, not the canonical business ledger.",
        "",
        "## Seed Audit",
        "",
    ]
    for record in data.get("seed_audit", []):
        dump_record_block(lines, record.get("标准名", "Untitled Seed Family"), record, AUDIT_HEADERS)
    lines.append("")
    return "\n".join(lines)


def parse_markdown_records(lines: List[str], section: str, title_key: str, ordered_keys: List[str], aliases: Dict[str, Iterable[str]]) -> List[Dict[str, str]]:
    records: List[Dict[str, str]] = []
    current: Dict[str, str] | None = None
    in_section = False
    for line in lines:
        stripped = line.strip()
        if stripped == f"## {section}":
            in_section = True
            current = None
            continue
        if in_section and stripped.startswith("## "):
            break
        if not in_section:
            continue
        if stripped.startswith("### "):
            if current:
                records.append(current)
            current = {title_key: stripped[4:].strip()}
            continue
        if current is not None and stripped.startswith("- "):
            body = stripped[2:]
            if ":" in body:
                key, value = body.split(":", 1)
                key = canonical_key(key.strip(), aliases)
                current[key] = value.strip()
    if current:
        records.append(current)
    normalized: List[Dict[str, str]] = []
    for record in records:
        normalized.append({key: sanitize(record.get(key, "")) for key in ordered_keys + ["Retrieval Hint"]})
    return normalized


def parse_markdown_summary(lines: List[str]) -> Dict[str, str]:
    summary: Dict[str, str] = {}
    in_section = False
    for line in lines:
        stripped = line.strip()
        if stripped == "## Summary":
            in_section = True
            continue
        if in_section and stripped.startswith("## "):
            break
        if in_section and stripped.startswith("- "):
            body = stripped[2:]
            if ":" in body:
                key, value = body.split(":", 1)
                key = canonical_key(key.strip(), SUMMARY_ALLOWED_KEYS)
                if key in SUMMARY_ALLOWED_KEYS:
                    summary[key] = value.strip()
    return summary


def parse_markdown_notes(lines: List[str]) -> List[Tuple[str, str]]:
    notes: List[Tuple[str, str]] = []
    in_section = False
    for line in lines:
        stripped = line.strip()
        if stripped == "## Search Notes":
            in_section = True
            continue
        if in_section and stripped.startswith("## "):
            break
        if in_section and stripped.startswith("- "):
            body = stripped[2:]
            if ":" in body:
                left, right = body.split(":", 1)
                notes.append((left.strip(), right.strip()))
            else:
                notes.append((body.strip(), ""))
    return notes


def read_existing_hints(md_path: Path) -> Dict[Tuple[str, str], str]:
    if not md_path.exists():
        return {}
    lines = md_path.read_text(encoding="utf-8").splitlines()
    hints: Dict[Tuple[str, str], str] = {}
    for section in ("Main Pool", "Tracking List"):
        current_title: str | None = None
        current_hint = ""
        in_section = False
        for line in lines:
            stripped = line.strip()
            if stripped == f"## {section}":
                in_section = True
                current_title = None
                current_hint = ""
                continue
            if in_section and stripped.startswith("## "):
                if current_title is not None:
                    hints[(section, current_title)] = current_hint
                break
            if not in_section:
                continue
            if stripped.startswith("### "):
                if current_title is not None:
                    hints[(section, current_title)] = current_hint
                current_title = stripped[4:].strip()
                current_hint = ""
                continue
            if current_title is not None and stripped.startswith("- Retrieval Hint:"):
                current_hint = stripped.split(":", 1)[1].strip()
        if in_section and current_title is not None:
            hints[(section, current_title)] = current_hint
    return hints


def markdown_to_data(md_path: Path, audit_md_path: Path | None = None) -> Dict[str, object]:
    lines = md_path.read_text(encoding="utf-8").splitlines()
    audit_lines = audit_md_path.read_text(encoding="utf-8").splitlines() if audit_md_path and audit_md_path.exists() else []
    return {
        "summary": parse_markdown_summary(lines),
        "main_pool": parse_markdown_records(lines, "Main Pool", "会议名称", MAIN_HEADERS, MAIN_HEADER_ALIASES),
        "tracking_list": parse_markdown_records(lines, "Tracking List", "会议名称", TRACKING_HEADERS, TRACKING_HEADER_ALIASES),
        "seed_audit": parse_markdown_records(audit_lines, "Seed Audit", "标准名", AUDIT_HEADERS, AUDIT_HEADER_ALIASES)
        if audit_lines
        else [],
        "search_notes": parse_markdown_notes(lines),
    }


def attach_hints(data: Dict[str, object], hint_map: Dict[Tuple[str, str], str]) -> Dict[str, object]:
    for section_name, key_name, data_key in [
        ("Main Pool", "会议名称", "main_pool"),
        ("Tracking List", "会议名称", "tracking_list"),
    ]:
        for record in data[data_key]:
            record["Retrieval Hint"] = hint_map.get((section_name, record[key_name]), "")
    return data


def parse_start_date_key(text: str) -> Tuple[int, date]:
    value = sanitize(text)
    if not value or value == "待确认":
        return (3, date.max)

    exact = re.search(r"(\d{4})-(\d{2})-(\d{2})", value)
    if exact:
        year, month, day = map(int, exact.groups())
        return (0, date(year, month, day))

    month_only = re.search(r"(\d{4})-(\d{2})(?!-\d{2})", value)
    if month_only:
        year, month = map(int, month_only.groups())
        return (1, date(year, month, 1))

    year_only = re.search(r"\b(\d{4})\b", value)
    if year_only:
        return (2, date(int(year_only.group(1)), 1, 1))

    return (3, date.max)


def sort_main_pool(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    return sorted(
        rows,
        key=lambda row: (
            parse_start_date_key(row.get("起止日期", ""))[0],
            parse_start_date_key(row.get("起止日期", ""))[1],
            sanitize(row.get("会议名称", "")),
        ),
    )


def style_sheet(ws) -> None:
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_summary_sheet(ws, summary: Dict[str, str]) -> None:
    ws.cell(1, 1, sanitize(summary.get("Report Title", "Conference Ledger Export")))
    row = 2
    for key in ["核验日期", "种子池家族总数", "本轮已核验", "确认国内会议", "持续跟踪"]:
        ws.cell(row, 1, f"{key}: {sanitize(summary.get(key, ''))}")
        row += 1
    ws.column_dimensions["A"].width = 80


def write_table_sheet(ws, headers: List[str], rows: List[Dict[str, str]]) -> None:
    ws.append(headers)
    for record in rows:
        ws.append([sanitize(record.get(h, "")) for h in headers])
    style_sheet(ws)
    for idx, header in enumerate(headers, start=1):
        width = max(len(header) + 4, 16)
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = min(width, 28)


def write_notes_sheet(ws, notes: List[Tuple[str, str]]) -> None:
    for left, right in notes:
        ws.append([left, right])
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 100
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def data_to_workbook(data: Dict[str, object], output_path: Path) -> None:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = SUMMARY_SHEET
    write_summary_sheet(summary_ws, data.get("summary", {}))

    main_rows = sort_main_pool(list(data.get("main_pool", [])))
    write_table_sheet(wb.create_sheet(MAIN_SHEET), MAIN_HEADERS, main_rows)
    write_table_sheet(wb.create_sheet(TRACKING_SHEET), TRACKING_HEADERS, list(data.get("tracking_list", [])))
    write_table_sheet(wb.create_sheet(AUDIT_SHEET), AUDIT_HEADERS, list(data.get("seed_audit", [])))
    write_notes_sheet(wb.create_sheet(NOTES_SHEET), list(data.get("search_notes", [])))
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync conference ledger between XLSX and Markdown.")
    sub = parser.add_subparsers(dest="cmd", required=True)

    xlsx_to_md = sub.add_parser("xlsx-to-md")
    xlsx_to_md.add_argument("xlsx")
    xlsx_to_md.add_argument("md")
    xlsx_to_md.add_argument("--audit-md", dest="audit_md")

    md_to_xlsx = sub.add_parser("md-to-xlsx")
    md_to_xlsx.add_argument("md")
    md_to_xlsx.add_argument("xlsx")
    md_to_xlsx.add_argument("--audit-md", dest="audit_md")

    args = parser.parse_args()
    if args.cmd == "xlsx-to-md":
        data = workbook_to_data(Path(args.xlsx))
        data = attach_hints(data, read_existing_hints(Path(args.md)))
        Path(args.md).write_text(data_to_markdown(data), encoding="utf-8")
        if args.audit_md:
            Path(args.audit_md).write_text(audit_to_markdown(data), encoding="utf-8")
    else:
        data = markdown_to_data(Path(args.md), Path(args.audit_md) if args.audit_md else None)
        data_to_workbook(data, Path(args.xlsx))


if __name__ == "__main__":
    main()
